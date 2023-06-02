import datetime as dt
import openpyxl as xl

start_time_val = "15:30"
end_time_val = "17:45"
#M.checked = True
options = ['Mon','Tue','Wed','Thur','Fri','Satuday','Sunday']
timeformat = '%H:%M'

SLpath = '/Users/oseongjae/PythonProject/NewProject/StudentList.xlsx'
AVpath = '/Users/oseongjae/PythonProject/NewProject/Availability.xlsx'
testWb = '/Users/oseongjae/PythonProject/NewProject/test.xlsx'
testlist = '/Users/oseongjae/PythonProject/NewProject/testlist.xlsx'

student_list = xl.load_workbook(SLpath)
availability_workbook = xl.load_workbook(AVpath)
test_wb = xl.load_workbook(testWb)
testlist_wb = xl.load_workbook(testlist)
selectedrow=5


def test():
    day = options[0] #Mon
    start_time = dt.datetime.strptime(start_time_val,timeformat)
    end_time = dt.datetime.strptime(end_time_val, timeformat)
    st_i = find_index('S', start_time)
    end_i = find_index('E', end_time)
    row = selectedrow
    ws = test_wb[day]
    ws.merge_cells(start_row=selectedrow, start_column=st_i, end_row=selectedrow, end_column=end_i)
    acell = ws.cell(row=selectedrow, column=st_i)
    acell.value = 'class {} - {}'.format(start_time_val, end_time_val)
    acell.alignment = xl.styles.Alignment(horizontal="center",vertical="center")
    acell.fill = xl.styles.PatternFill("solid",bgColor="FFFF54",fgColor="FFFF54")
    test_wb.save(testWb)



def find_index(SorE, time):
    OG_time = "7:00"
    #time = dt.datetime.strptime(time, timeformat)
    OG_start = dt.datetime.strptime(OG_time, timeformat)
    start_time_gap = time - OG_start
    v = start_time_gap.total_seconds()
    i = int(int(v/60)/15)

    if SorE == 'S':
        return i+2
    if SorE == 'E':
        return i+1


def add_class():
    for day in options:
        if day.checked == True:
            start_time = dt.datetime.strptime(start_time_val,timeformat)
            end_time = dt.datetime.strptime(end_time_val, timeformat)
            st_i = find_index('S', start_time)
            end_i = find_index('E', end_time)
            row = selectedrow
            ws = availability_workbook[day]
            ws.merge_cells(start_row=selectedrow, start_column=st_i, end_row=selectedrow, end_column=end_i)
            acell = ws.cell(row=selectedrow, column=st_i)
            acell.value = 'class {} - {}'.format(start_time_val, end_time_val)
            acell.alignment = xl.styles.Alignment(horizontal="center",vertical="center")
            acell.fill = xl.styles.PatternFill("solid",bgColor="00FFFF00")


test()
